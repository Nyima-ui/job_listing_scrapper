import csv
from openpyxl import load_workbook


def write_to_excel():

    # Load existing Excel workbook
    wb = load_workbook("jobs_data.xlsx")
    ws = wb.active  # or wb["SheetName"] if you want a specific sheet

    # Clear existing data
    for row in ws.iter_rows(
        min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
    ):
        for cell in row:
            cell.value = None

    # Open CSV and write rows to Excel (overwrite old data)
    with open("jobs_data.csv", newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        for row_idx, row in enumerate(reader, start=1):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    # Save the workbook (overwrites the same file)
    wb.save("jobs_data.xlsx")

if __name__ == "__main__": 
    write_to_excel()